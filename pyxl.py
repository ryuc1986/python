import openpyxl

book = openpyxl.Workbook()
active_sheet = book.active
active_sheet['A1'] = 10
active_sheet['B2'] = 20
active_sheet['C3'] = 30

book.create_sheet('datasheet')
book.save('data.xlsx')


wb = openpyxl.load_workbook('data.xlsx')

